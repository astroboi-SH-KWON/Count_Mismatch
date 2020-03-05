import pandas as pd
import openpyxl
from time import clock
"""
length가 같은 training과 test의 seq를 idx순서로 1:1 비교하여
test 기준으로 training에 mismatch 되는 갯수를 세어준다
result 는 test seq 목록에 각각mismatch 갯수
"""
############### 환경 설정 시작 ################
# mismatch 갯수
CNT = 2 # 0, 1, 2개 각각 확인
# CNT = 3 # 0, 1, 2, 3 개 각각 확인

# 엑셀 파일 경로
FILE_PATH = "D:/000_WORK/KimNahye/20200304/Lib 10fg - 복사본.xlsx"
RESULT_PATH = "D:/000_WORK/KimNahye/20200304/Lib 10fg_result.xlsx"
# FILE_PATH = "D:/000_WORK/KimNahye/20200304/kwon 쌤.xlsx"
# RESULT_PATH = "D:/000_WORK/KimNahye/20200304/kwon 쌤_result.xlsx"

# 각 sheet 별로 seq data 가져 온다
TRAINING_SET = pd.read_excel(FILE_PATH , sheet_name='training')
TEST_SET = pd.read_excel(FILE_PATH , sheet_name='test') # 기준

# Target 컬럼을 가져 오는 경우
# training_targets = TRAINING_SET['Target']
# test_targets = TEST_SET['Target']
training_targets = TRAINING_SET['Guide (X19)']
test_targets = TEST_SET['guide']

# Target context sequence 컬럼을 가져 오는 경우
# training_targets = TRAINING_SET['Target context sequence']
# test_targets = TEST_SET['Target context sequence']
############### 환경 설정 끝 ################

# df = pd.DataFrame(excel_file)
# targets = df['Target']

# mismatch 찾기
def mismatch(test_target,training_target):
    count = 0
    for i in range(0,len(test_target)):
        if test_target[i] != training_target[i]:
            count +=1
        # count 값이 CNT 값을 초과하면 loop 탈출
        if count > CNT: break
    return count

# 결과 print만 하는 함수
def get_data(tests,trainings):
    for test in tests:
        if pd.notnull(test):
            for training in trainings:
                if pd.notnull(training):
                    if mismatch(test, training) <= CNT:
                        print(test + "\n" + training + " >>>", mismatch(test, training))

# 중복 seq 제외 하기 위한 filter
def make_set(targets):
    return {x for x in targets if pd.notna(x)}
    # return set(sets)

# 각 seq 마다 mismatch 담기 위한 dict 
def make_dic(set_val):
    seq_dict = dict()
    for s in set_val:
        mismtch_dict = dict()
        for tmp in range(0,CNT+1):
            mismtch_dict[str(tmp)]=0
        seq_dict[s] = mismtch_dict
    return seq_dict

# 결과 생성
def get_data2(tests_dict,trainings):
    for test in tests_dict.keys():
        if pd.notnull(test):
            for training in trainings:
                if pd.notnull(training):
                    if mismatch(test, training) <= CNT:
                        temp = tests_dict[test]
                        temp[str(mismatch(test, training) )] += 1
    return tests_dict

# get_data의 프린트 결과를 excel로
def get_data3(tests,trainings):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    row = 1
    sheet.cell(row=row, column=1, value="guide")
    sheet.cell(row=row, column=2, value="Guide (X19)")
    sheet.cell(row=row, column=3, value="EA")
    row = 2
    for test in tests:
        if pd.notnull(test):
            for training in trainings:
                if pd.notnull(training):
                    cnt_result = mismatch(test, training)
                    if cnt_result <= CNT:
                        sheet.cell(row=row, column=1, value=test)
                        sheet.cell(row=row, column=2, value=training)
                        sheet.cell(row=row, column=3, value=cnt_result)
                        row = row + 1
    workbook.save(filename=RESULT_PATH)

# 결과를 excel로 만드는 함수
def make_excel(result):
    df = pd.DataFrame(result)
    df_transe = df.T
    df_transe.to_excel(RESULT_PATH)


def main():
    get_data3(test_targets, training_targets)

    # target_set = make_set(test_targets)
    # target_dict = make_dic(target_set)
    # result_dict = get_data2(target_dict,training_targets)
    # print(result_dict)
    # make_excel(result_dict)



start_time = clock()
main()
print("::::::::::: %.2f seconds ::::::::::::::" % (clock() - start_time))