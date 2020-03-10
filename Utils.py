import pandas as pd
import openpyxl
from time import clock

class Utils:
    def __init__(self, prefix):
        self.path = prefix[0]
        self.result = prefix[1]
        self.sheet_name = prefix[2]
        self.col_name = prefix[3]
        self.ext_xlsx = ".xlsx"

    """
    get_excel : read excel file by sheet_name then make data to dict
    :param
        path :
        sheet_name :
    :return
        dict object : { 'col_name' : {0: 'val0' , 1: 'val1' , 2 : 'val2' ...
    """
    def get_excel(self):
        return pd.read_excel(self.path , sheet_name=self.sheet_name).to_dict()

    """
    get_data_by_col : get data by key(=col_name)
    :return 
        list object : [ ]
    """
    def get_data_by_col(self, tmp_dict):
        return tmp_dict[self.col_name].values()

    """
    make_set : filtering reduplication
    """
    def make_set(self,tmp_dict):
        return {x for x in tmp_dict if pd.notna(x)}

    """
    make_dic : 
    :return
        seq_dict : dict object 
                {
                    key : {0: 0 , 1: 0 , 2 : 0 ... depending on cnt
                }
    """
    def make_dic(self,keys, cnt):
        seq_dict = dict()
        for s in keys:
            mismtch_dict = dict()
            for tmp in range(0, cnt + 1):
                mismtch_dict[str(tmp)] = 0
            seq_dict[s] = mismtch_dict
        return seq_dict

    """
    make_excel_by_pandas : make excel file with pandas.DataFrame object
    """
    def make_excel_by_pandas(self):
        df = pd.DataFrame(self.result)
        df_transe = df.T
        df_transe.to_excel(self.path + str(clock()) + self.ext_xlsx)

    """
    make_excel_by_openpyxl : make excel file with dict object by openpyxl
    :param
        col_arr : column names
        self.result : dict object { guide : [ guide , cnt ] }
    """
    def make_excel_by_openpyxl(self, col_arr):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        row = 1
        sheet.cell(row=row, column=1, value=col_arr[0])
        sheet.cell(row=row, column=2, value=col_arr[1])
        sheet.cell(row=row, column=3, value=col_arr[2])
        row = 2
        for key, val_arr in self.result.items():
            col = 1
            sheet.cell(row=row, column=col, value=key)
            for val in val_arr:
                col = col + 1
                sheet.cell(row=row, column=col, value=val)
            row = row + 1
        workbook.save(filename=self.path + str(clock()) + self.ext_xlsx)
